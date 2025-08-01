#!/usr/bin/env python3
"""
KYC Sentinel - Desktop Application for KYC/AML Screening
Uses OpenSanctions API for compliance checking
"""

import tkinter as tk
from tkinter import ttk, filedialog, messagebox
import pandas as pd
import requests
import openpyxl
from openpyxl.styles import PatternFill
import os
import sys
from datetime import datetime
import threading
import json

class KYCSentinel:
    def __init__(self, root):
        self.root = root
        self.root.title("KYC Sentinel - Compliance Screening Tool")
        self.root.geometry("600x500")
        self.root.resizable(True, True)
        
        # Configure style
        style = ttk.Style()
        style.theme_use('clam')
        
        self.setup_ui()
        
    def setup_ui(self):
        """Setup the main user interface"""
        # Main frame
        main_frame = ttk.Frame(self.root, padding="20")
        main_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        
        # Configure grid weights
        self.root.columnconfigure(0, weight=1)
        self.root.rowconfigure(0, weight=1)
        main_frame.columnconfigure(1, weight=1)
        
        # Title
        title_label = ttk.Label(main_frame, text="KYC Sentinel", 
                               font=("Arial", 16, "bold"))
        title_label.grid(row=0, column=0, columnspan=2, pady=(0, 20))
        
        # Mode selection
        mode_frame = ttk.LabelFrame(main_frame, text="Select Mode", padding="10")
        mode_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        mode_frame.columnconfigure(0, weight=1)
        mode_frame.columnconfigure(1, weight=1)
        
        # Manual mode button
        self.manual_btn = ttk.Button(mode_frame, text="Manual Mode", 
                                   command=self.show_manual_mode)
        self.manual_btn.grid(row=0, column=0, padx=(0, 10), sticky=(tk.W, tk.E))
        
        # Batch mode button
        self.batch_btn = ttk.Button(mode_frame, text="Batch Mode (CSV)", 
                                  command=self.show_batch_mode)
        self.batch_btn.grid(row=0, column=1, padx=(10, 0), sticky=(tk.W, tk.E))
        
        # Content frame (will switch between manual and batch)
        self.content_frame = ttk.Frame(main_frame)
        self.content_frame.grid(row=2, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S))
        self.content_frame.columnconfigure(1, weight=1)
        main_frame.rowconfigure(2, weight=1)
        
        # Progress bar (initially hidden)
        self.progress_var = tk.StringVar()
        self.progress_label = ttk.Label(main_frame, textvariable=self.progress_var)
        self.progress_label.grid(row=3, column=0, columnspan=2, pady=(10, 0))
        
        self.progress_bar = ttk.Progressbar(main_frame, mode='indeterminate')
        self.progress_bar.grid(row=4, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(5, 0))
        
        # Hide progress initially
        self.progress_label.grid_remove()
        self.progress_bar.grid_remove()
        
        # Show manual mode by default
        self.show_manual_mode()
        
    def show_manual_mode(self):
        """Display manual input mode"""
        # Clear content frame
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # Manual input form
        form_frame = ttk.LabelFrame(self.content_frame, text="Manual Entry", padding="15")
        form_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E), pady=(0, 20))
        form_frame.columnconfigure(1, weight=1)
        
        # Name field
        ttk.Label(form_frame, text="Name:").grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
        self.name_var = tk.StringVar()
        name_entry = ttk.Entry(form_frame, textvariable=self.name_var, width=40)
        name_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 10), padx=(10, 0))
        
        # DOB field
        ttk.Label(form_frame, text="Date of Birth:").grid(row=1, column=0, sticky=tk.W, pady=(0, 10))
        self.dob_var = tk.StringVar()
        dob_entry = ttk.Entry(form_frame, textvariable=self.dob_var, width=40)
        dob_entry.grid(row=1, column=1, sticky=(tk.W, tk.E), pady=(0, 10), padx=(10, 0))
        ttk.Label(form_frame, text="(Format: YYYY-MM-DD)", font=("Arial", 8)).grid(row=1, column=2, sticky=tk.W, padx=(5, 0))
        
        # CNIC field
        ttk.Label(form_frame, text="CNIC:").grid(row=2, column=0, sticky=tk.W, pady=(0, 10))
        self.cnic_var = tk.StringVar()
        cnic_entry = ttk.Entry(form_frame, textvariable=self.cnic_var, width=40)
        cnic_entry.grid(row=2, column=1, sticky=(tk.W, tk.E), pady=(0, 10), padx=(10, 0))
        
        # Run check button
        check_btn = ttk.Button(form_frame, text="Run KYC Check", 
                             command=self.run_manual_check, 
                             style='Accent.TButton')
        check_btn.grid(row=3, column=0, columnspan=2, pady=(20, 0))
        
        # Results display
        results_frame = ttk.LabelFrame(self.content_frame, text="Results", padding="15")
        results_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(10, 0))
        results_frame.columnconfigure(0, weight=1)
        self.content_frame.rowconfigure(1, weight=1)
        
        # Results text widget with scrollbar
        text_frame = ttk.Frame(results_frame)
        text_frame.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        text_frame.columnconfigure(0, weight=1)
        text_frame.rowconfigure(0, weight=1)
        results_frame.rowconfigure(0, weight=1)
        
        self.results_text = tk.Text(text_frame, height=10, wrap=tk.WORD)
        scrollbar = ttk.Scrollbar(text_frame, orient=tk.VERTICAL, command=self.results_text.yview)
        self.results_text.configure(yscrollcommand=scrollbar.set)
        
        self.results_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        
    def show_batch_mode(self):
        """Display batch processing mode"""
        # Clear content frame
        for widget in self.content_frame.winfo_children():
            widget.destroy()
            
        # Batch processing form
        batch_frame = ttk.LabelFrame(self.content_frame, text="Batch Processing", padding="15")
        batch_frame.grid(row=0, column=0, columnspan=2, sticky=(tk.W, tk.E))
        batch_frame.columnconfigure(1, weight=1)
        
        # File selection
        ttk.Label(batch_frame, text="CSV File:").grid(row=0, column=0, sticky=tk.W, pady=(0, 10))
        self.file_var = tk.StringVar()
        file_entry = ttk.Entry(batch_frame, textvariable=self.file_var, width=50)
        file_entry.grid(row=0, column=1, sticky=(tk.W, tk.E), pady=(0, 10), padx=(10, 5))
        
        browse_btn = ttk.Button(batch_frame, text="Browse", command=self.browse_file)
        browse_btn.grid(row=0, column=2, pady=(0, 10), padx=(5, 0))
        
        # Instructions
        instructions = ttk.Label(batch_frame, 
                               text="CSV should have columns: Name, DOB, CNIC", 
                               font=("Arial", 9))
        instructions.grid(row=1, column=0, columnspan=3, pady=(0, 15))
        
        # Process button
        process_btn = ttk.Button(batch_frame, text="Upload CSV & Run Bulk Check", 
                               command=self.run_batch_check,
                               style='Accent.TButton')
        process_btn.grid(row=2, column=0, columnspan=3, pady=(10, 0))
        
        # Results summary
        summary_frame = ttk.LabelFrame(self.content_frame, text="Processing Summary", padding="15")
        summary_frame.grid(row=1, column=0, columnspan=2, sticky=(tk.W, tk.E, tk.N, tk.S), pady=(20, 0))
        summary_frame.columnconfigure(0, weight=1)
        self.content_frame.rowconfigure(1, weight=1)
        
        self.summary_text = tk.Text(summary_frame, height=8, wrap=tk.WORD)
        summary_scrollbar = ttk.Scrollbar(summary_frame, orient=tk.VERTICAL, command=self.summary_text.yview)
        self.summary_text.configure(yscrollcommand=summary_scrollbar.set)
        
        self.summary_text.grid(row=0, column=0, sticky=(tk.W, tk.E, tk.N, tk.S))
        summary_scrollbar.grid(row=0, column=1, sticky=(tk.N, tk.S))
        summary_frame.rowconfigure(0, weight=1)
        
    def browse_file(self):
        """Open file dialog to select CSV file"""
        filename = filedialog.askopenfilename(
            title="Select CSV File",
            filetypes=[("CSV files", "*.csv"), ("All files", "*.*")],
            initialdir=os.path.join(os.path.dirname(__file__), "input")
        )
        if filename:
            self.file_var.set(filename)
            
    def check_sanctions(self, name):
        """Check name against OpenSanctions API"""
        try:
            # API endpoint
            url = "https://api.opensanctions.org/match"
            params = {'q': name}
            
            # Make request with timeout
            response = requests.get(url, params=params, timeout=10)
            response.raise_for_status()
            
            data = response.json()
            
            # Check if matches found
            if 'matches' in data and len(data['matches']) > 0:
                return "Match Found", data['matches']
            else:
                return "Clear", []
                
        except requests.exceptions.RequestException as e:
            raise Exception(f"API request failed: {str(e)}")
        except json.JSONDecodeError:
            raise Exception("Invalid response from API")
        except Exception as e:
            raise Exception(f"Unexpected error: {str(e)}")
            
    def run_manual_check(self):
        """Run KYC check for manual entry"""
        name = self.name_var.get().strip()
        dob = self.dob_var.get().strip()
        cnic = self.cnic_var.get().strip()
        
        if not name:
            messagebox.showerror("Error", "Name is required")
            return
            
        # Clear previous results
        self.results_text.delete(1.0, tk.END)
        
        # Show progress
        self.show_progress("Checking sanctions database...")
        
        def check_thread():
            try:
                status, matches = self.check_sanctions(name)
                
                # Prepare results
                result_text = f"Name: {name}\n"
                result_text += f"DOB: {dob}\n"
                result_text += f"CNIC: {cnic}\n"
                result_text += f"Status: {status}\n"
                result_text += f"Checked on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}\n\n"
                
                if status == "Match Found":
                    result_text += "⚠️ MATCHES FOUND:\n"
                    for i, match in enumerate(matches[:3], 1):  # Show first 3 matches
                        result_text += f"\n{i}. {match.get('name', 'N/A')}\n"
                        result_text += f"   Score: {match.get('score', 'N/A')}\n"
                        result_text += f"   Dataset: {match.get('dataset', 'N/A')}\n"
                else:
                    result_text += "✅ No sanctions matches found.\n"
                
                # Update UI in main thread
                self.root.after(0, lambda: self.update_manual_results(result_text, name, dob, cnic, status))
                
            except Exception as e:
                error_msg = f"Error during check: {str(e)}"
                self.root.after(0, lambda: self.show_error(error_msg))
            finally:
                self.root.after(0, self.hide_progress)
                
        # Start check in background thread
        threading.Thread(target=check_thread, daemon=True).start()
        
    def update_manual_results(self, result_text, name, dob, cnic, status):
        """Update manual results display"""
        self.results_text.delete(1.0, tk.END)
        self.results_text.insert(1.0, result_text)
        
        # Save to Excel
        try:
            self.save_to_excel([{
                'Name': name,
                'DOB': dob,
                'CNIC': cnic,
                'Status': status
            }])
            self.results_text.insert(tk.END, f"\n📄 Results saved to: {os.path.join('output', 'KYC_Results.xlsx')}")
        except Exception as e:
            self.results_text.insert(tk.END, f"\n❌ Error saving Excel: {str(e)}")
            
    def run_batch_check(self):
        """Run KYC check for batch CSV file"""
        csv_file = self.file_var.get().strip()
        
        if not csv_file:
            messagebox.showerror("Error", "Please select a CSV file")
            return
            
        if not os.path.exists(csv_file):
            messagebox.showerror("Error", "Selected file does not exist")
            return
            
        # Clear previous results
        self.summary_text.delete(1.0, tk.END)
        
        # Show progress
        self.show_progress("Processing CSV file...")
        
        def batch_thread():
            try:
                # Read CSV
                df = pd.read_csv(csv_file)
                
                # Validate columns
                required_cols = ['Name', 'DOB', 'CNIC']
                missing_cols = [col for col in required_cols if col not in df.columns]
                if missing_cols:
                    raise Exception(f"Missing required columns: {', '.join(missing_cols)}")
                
                results = []
                total_rows = len(df)
                
                self.root.after(0, lambda: self.update_progress(f"Processing {total_rows} records..."))
                
                for index, row in df.iterrows():
                    name = str(row.get('Name', '')).strip()
                    dob = str(row.get('DOB', '')).strip()
                    cnic = str(row.get('CNIC', '')).strip()
                    
                    if not name:
                        status = "Error: No name provided"
                    else:
                        try:
                            status, _ = self.check_sanctions(name)
                        except Exception as e:
                            status = f"Error: {str(e)}"
                    
                    results.append({
                        'Name': name,
                        'DOB': dob,
                        'CNIC': cnic,
                        'Status': status
                    })
                    
                    # Update progress
                    progress_text = f"Processing {index + 1}/{total_rows}: {name}"
                    self.root.after(0, lambda text=progress_text: self.update_progress(text))
                
                # Save results
                self.save_to_excel(results)
                
                # Generate summary
                total_checked = len(results)
                matches_found = len([r for r in results if r['Status'] == 'Match Found'])
                clear_results = len([r for r in results if r['Status'] == 'Clear'])
                errors = len([r for r in results if r['Status'].startswith('Error')])
                
                summary = f"Batch Processing Complete!\n\n"
                summary += f"Total Records: {total_checked}\n"
                summary += f"✅ Clear: {clear_results}\n"
                summary += f"⚠️ Matches Found: {matches_found}\n"
                summary += f"❌ Errors: {errors}\n\n"
                summary += f"Results saved to: output/KYC_Results.xlsx\n"
                summary += f"Completed on: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
                
                self.root.after(0, lambda: self.update_batch_summary(summary))
                
            except Exception as e:
                error_msg = f"Batch processing error: {str(e)}"
                self.root.after(0, lambda: self.show_error(error_msg))
            finally:
                self.root.after(0, self.hide_progress)
                
        # Start batch processing in background thread
        threading.Thread(target=batch_thread, daemon=True).start()
        
    def update_batch_summary(self, summary):
        """Update batch processing summary"""
        self.summary_text.delete(1.0, tk.END)
        self.summary_text.insert(1.0, summary)
        
    def save_to_excel(self, results):
        """Save results to Excel file with styling"""
        # Ensure output directory exists
        output_dir = os.path.join(os.path.dirname(__file__), "output")
        os.makedirs(output_dir, exist_ok=True)
        
        # Create Excel file
        output_file = os.path.join(output_dir, "KYC_Results.xlsx")
        
        # Create workbook and worksheet
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "KYC Results"
        
        # Headers
        headers = ['Name', 'DOB', 'CNIC', 'Status']
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header)
            cell.font = openpyxl.styles.Font(bold=True)
            
        # Data rows
        red_fill = PatternFill(start_color='FFCCCC', end_color='FFCCCC', fill_type='solid')
        
        for row_idx, result in enumerate(results, 2):
            ws.cell(row=row_idx, column=1, value=result['Name'])
            ws.cell(row=row_idx, column=2, value=result['DOB'])
            ws.cell(row=row_idx, column=3, value=result['CNIC'])
            status_cell = ws.cell(row=row_idx, column=4, value=result['Status'])
            
            # Highlight matches in red
            if result['Status'] == 'Match Found':
                for col in range(1, 5):
                    ws.cell(row=row_idx, column=col).fill = red_fill
                    
        # Auto-adjust column widths
        for column in ws.columns:
            max_length = 0
            column_letter = column[0].column_letter
            for cell in column:
                try:
                    if len(str(cell.value)) > max_length:
                        max_length = len(str(cell.value))
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column_letter].width = adjusted_width
            
        # Save file
        wb.save(output_file)
        
    def show_progress(self, message):
        """Show progress bar and message"""
        self.progress_var.set(message)
        self.progress_label.grid()
        self.progress_bar.grid()
        self.progress_bar.start(10)
        
    def update_progress(self, message):
        """Update progress message"""
        self.progress_var.set(message)
        
    def hide_progress(self):
        """Hide progress bar"""
        self.progress_bar.stop()
        self.progress_label.grid_remove()
        self.progress_bar.grid_remove()
        
    def show_error(self, message):
        """Show error message"""
        messagebox.showerror("Error", message)
        if hasattr(self, 'results_text'):
            self.results_text.delete(1.0, tk.END)
            self.results_text.insert(1.0, f"❌ {message}")
        if hasattr(self, 'summary_text'):
            self.summary_text.delete(1.0, tk.END)
            self.summary_text.insert(1.0, f"❌ {message}")

def main():
    """Main application entry point"""
    root = tk.Tk()
    app = KYCSentinel(root)
    
    # Center window on screen
    root.update_idletasks()
    x = (root.winfo_screenwidth() // 2) - (root.winfo_width() // 2)
    y = (root.winfo_screenheight() // 2) - (root.winfo_height() // 2)
    root.geometry(f"+{x}+{y}")
    
    root.mainloop()

if __name__ == "__main__":
    main()